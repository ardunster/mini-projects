#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 11 17:43:42 2020

@author: adunster
"""
'''
This program is designed to be run once on data procured from https://www.gold.org/download/8369
to provide a baseline file to run goldsilver.py with.
This data is available only to registered users and is not allowed for redistribution.
However, registration is free and this script will extract the preceding 400 
business days from the file, as provided on 04/11/20, and calculate 200 dma, 
and ratio of current price to 200dma, from the bottom 200 entries. If their 
formatting changes the script may not work correctly. Double check values 
before proceeding.
'''

import openpyxl as xl
from os import path


prices_big = path.expanduser('~') + '/Downloads/Prices.xlsx'

wb_import = xl.load_workbook(prices_big)
ws_import = wb_import.worksheets[9]

if not path.exists('goldprices.xlsx'):
    prices_setup = xl.Workbook()
    sheet_setup = prices_setup.active
    sheet_setup['A1'] = 'Gold prices by day'
    sheet_setup['A2'] = 'Date'
    sheet_setup['B2'] = 'Price (USD)'
    sheet_setup['C2'] = 'Change from previous'
    sheet_setup['D2'] = '200dma'
    sheet_setup['E2'] = 'Ratio'
    prices_setup.save('goldprices.xlsx')


prices_xlsx = 'goldprices.xlsx'
prices_wb = xl.load_workbook(prices_xlsx)
prices_ws = prices_wb.worksheets[0]

prices_mr = prices_ws.max_row
prices_mc = prices_ws.max_column


if not ws_import.title == 'Daily':
    print('Wrong sheet, check data')
else:
    import_mr = ws_import.max_row
    for row in range(import_mr,0,-1):
        if not ws_import.cell(row=row,column=4).value == None:
            last_row = row
            first_row = (row - 400)
            break


add_row = (int(prices_mr) + 1)
count = 1

for i in range(first_row,last_row):
    w_date = ws_import.cell(row=i,column=4).value
    w_date = w_date.date() # convert from datetime to date
    prices_ws.cell(row=add_row, column=1).value = w_date
    prices_ws.cell(row=add_row, column=1).number_format = 'yy/mm/dd'
    prices_ws.cell(row=add_row, column=2).value = ws_import.cell(row=i,column=5).value
    prices_ws.cell(row=add_row, column=2).number_format = '#,##0.00'
    if count > 1:
        prices_ws.cell(row=add_row, column=3).value = '=(B{}-B{})'.format(add_row,add_row-1)
        prices_ws.cell(row=add_row, column=3).number_format = '#,##0.00'
    if count >= 200:
        prices_ws.cell(row=add_row, column=4).value = '=AVERAGE(B{}:B{})'.format(add_row,add_row-199)
        prices_ws.cell(row=add_row, column=4).number_format = '#,##0.00'
        prices_ws.cell(row=add_row, column=5).value = '=(B{}/D{})'.format(add_row, add_row)
        prices_ws.cell(row=add_row, column=5).number_format = '#,##0.00'
    add_row += 1
    count += 1

prices_wb.save('goldprices.xlsx')

