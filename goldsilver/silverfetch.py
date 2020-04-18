#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 15 14:58:17 2020

@author: adunster
"""
'''
The first section of this script fetches current year price history from 
https://goldsilver.com/price-charts/historical-london-fix/ 
Prior year price history is also available, but, it displays on the page via a 
script without changing the address or source. I spent some time looking for a 
way to reliably access it via code with no luck, and decided for the purposes of 
this program it was easier and more efficient to copy and paste the entire table 
into a text document and work with that.

The rest of the script then produces an xlsx file from this data as history 
for the goldsilver.py script. 
'''

from bs4 import BeautifulSoup as bs
import requests
import openpyxl as xl
from os import path
import datetime


'''
Date and price into tuples
'''


url = 'https://goldsilver.com/price-charts/historical-london-fix/'

def get_content(url):
    
    req = requests.get(url)
    soup = bs(req.content, 'html.parser')
    content = []

    for item in soup.find_all('td'):
        item_text = item.get_text()
        if not item_text == '':
            content.append(item_text)
        
    return content

y2020 = get_content(url)



y2020_trim = []

for i in range(0,(len(y2020) - 3),4):
    y2020_trim.append((y2020[i],y2020[i+3].strip('$')))

y2020_trim = y2020_trim[::-1]


with open('silver2019.txt') as text_y2019:
    y2019 = [line.split('\t') for line in text_y2019]

y2019_trim = [] 

for i in range(len(y2019)):
    y2019_trim.append((y2019[i][0],y2019[i][3].strip('$\n')))

y2019_trim = y2019_trim[::-1]


with open('silver2018.txt') as text_y2018:
    y2018 = [line.split('\t') for line in text_y2018]

y2018_trim = [] 

for i in range(len(y2018)):
    y2018_trim.append((y2018[i][0],y2018[i][3].strip('$\n')))

y2018_trim = y2018_trim[::-1]
    
    
'''
Tuples into xlsx file
'''

if not path.exists('silverprices.xlsx'):
    prices_setup = xl.Workbook()
    sheet_setup = prices_setup.active
    sheet_setup['A1'] = 'Silver prices by day'
    sheet_setup['A2'] = 'Date'
    sheet_setup['B2'] = 'Price (USD)'
    sheet_setup['C2'] = 'Change from previous'
    sheet_setup['D2'] = '200dma'
    sheet_setup['E2'] = 'Ratio'
    prices_setup.save('silverprices.xlsx')


prices_xlsx = 'silverprices.xlsx'
prices_wb = xl.load_workbook(prices_xlsx)
prices_ws = prices_wb.worksheets[0]

prices_mr = prices_ws.max_row
prices_mc = prices_ws.max_column


def tuples_to_xlsx(lst):
    '''
    Input: list of tuples of (date as YYYY-MM-DD, silver price as float)
    '''
    
    global add_row
    global count
    

    count = 1
    
    for i in range(len(lst)):
        if not lst[i][1] == '-':
            w_date = datetime.datetime.strptime(lst[i][0], '%Y-%m-%d')
            prices_ws.cell(row=add_row, column=1).value = w_date
            prices_ws.cell(row=add_row, column=1).number_format = 'yy/mm/dd'
            prices_ws.cell(row=add_row, column=2).value = float(lst[i][1])
            prices_ws.cell(row=add_row, column=2).number_format = '#,##0.00'
            if count > 1:
                prices_ws.cell(row=add_row, column=3).value = '=(B{}-B{})'.format(add_row,add_row-1)
                prices_ws.cell(row=add_row, column=3).number_format = '#,##0.00'
            if add_row > 200:
                prices_ws.cell(row=add_row, column=4).value = '=AVERAGE(B{}:B{})'.format(add_row,add_row-200)
                prices_ws.cell(row=add_row, column=4).number_format = '#,##0.00'
                prices_ws.cell(row=add_row, column=5).value = '=(B{}/D{})'.format(add_row, add_row)
                prices_ws.cell(row=add_row, column=5).number_format = '#,##0.00'
            add_row += 1
        count += 1
    

add_row = (int(prices_mr) + 1)

tuples_to_xlsx(y2018_trim)
tuples_to_xlsx(y2019_trim)
tuples_to_xlsx(y2020_trim)

prices_wb.save('silverprices.xlsx')




