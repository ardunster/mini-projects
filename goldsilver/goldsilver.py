#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 18:43:45 2020

@author: adunster
"""

'''
Gold/Silver
This script is designed to fetch the London Fix gold and silver prices daily from  
https://goldsilver.com/price-charts/historical-london-fix/

It then calculates the 200 day moving average and relative ratio of current 
price to the 200 dma for each, and raises and emails an alert if the relative 
ratio falls in either a high or low zone. 

In order to properly function the script depends on an xlsx file containing a 
minimum of 200 days of price data (properly formatted). I wrote two additional 
scripts to procure this data, first gold off of the first dataset I located 
(extractgoldprices.py) and later silver data off the above mentioned website 
(silverfetch.py). Either of those scripts will generate an empty xlsx file in 
the appropriate format (and filename) for this script to use, but data is not 
included with this upload due to terms of service and similar restrictions.

Also in this repository are an email script that can function off either 
localhost (in which case you will need a terminal running SMTP to catch any 
alerts) or an actual email server (run send_email.py directly to configure, or 
manually edit email_config.txt), and a set of tests to ensure functions are 
working as intended.
#some items may not yet be complete and included - but are planned before I 
#call it "done"

'''

# Imports

import openpyxl as xl
import datetime
import requests
import re
import send_email
from os import path
from bs4 import BeautifulSoup as bs



# Classes

class TooFewLines(Exception):
    '''
    Raised by too few lines in xlsx files
    '''
    pass

class MissingFiles(Exception):
    '''
    Raised if missing xlsx files
    '''
    pass


#Functions

def load_xlsx():
    '''
    Loads existing data
    '''
    global prices_gold
    global gold_wb
    global gold_ws 
    global gold_mr
    
    global prices_silver
    global silver_wb
    global silver_ws
    global silver_mr
    
    gold_wb = xl.load_workbook(prices_gold)
    gold_ws = gold_wb.worksheets[0]
    gold_mr = gold_ws.max_row
    
    silver_wb = xl.load_workbook(prices_silver)
    silver_ws = silver_wb.worksheets[0]
    silver_mr = silver_ws.max_row
    
    if gold_mr < 202:
        raise TooFewLines('Too few lines in {} to compute 200dma, check data.'.format(prices_gold))
    elif silver_mr < 202:
        raise TooFewLines('Too few lines in {} to compute 200dma, check data.'.format(prices_silver))


def get_current():
    '''
    Retrieves current information from website, returns as a list of tuples:
    (Datetime, Gold AM, Gold PM, Silver noon)
    '''
    
    url = 'https://goldsilver.com/price-charts/historical-london-fix/'
    
    req = requests.get(url)
    soup = bs(req.content, 'html.parser')
    content = []

    for item in soup.find_all('td'):
        item_text = item.get_text()
        if not item_text == '':
            content.append(item_text)
    
    content_arr = []
    re_sub = re.compile('[^\d\.\-]')
    
    for i in range(0,(len(content) - 3),4):
        w_date = datetime.datetime.strptime(content[i], '%Y-%m-%d')
        w_date = w_date.date()
        content_arr.append((w_date,(re.sub(re_sub, '', content[i+1])),(re.sub(re_sub, '', content[i+2])),(re.sub(re_sub, '', content[i+3]))))

    return content_arr


def verify_mr(goldorsilver):
    '''
    Checks against a possible bug in openpyxl reporting too high of a max row.
    Input: string, either 'gold' or 'silver'
    Output: verified last row (int)
    Starts with reported last row and works backward to verify that the contents
    of the cell is not None
    '''
    
    if goldorsilver == 'gold':
        verify_mr = gold_mr
        verify_ws = gold_ws
    elif goldorsilver == 'silver':
        verify_mr = silver_mr
        verify_ws = silver_ws
    for row in range(verify_mr,0,-1):
        if not verify_ws.cell(row=row,column=1).value == None:
            last_row = row
            return last_row


def update():
    '''
    Finds indexes for last updated information in both files, adds data after 
    index, saves files
    '''
    
    def find_index():
        ''' Checks indexes to find next line to add'''
        w_date_gold = gold_ws.cell(row=gold_mr, column=1).value.date()
        w_date_silver = silver_ws.cell(row=silver_mr, column=1).value.date()
        ind_gold = ''
        ind_silver = ''
        while ind_gold == '' and ind_silver == '':
            for tup in range(len(current)):
                if current[tup][0] == w_date_gold:
                    ind_gold = (tup - 1)
                elif current[tup][0].timetuple().tm_mon == 1 and len(current) == 1:
                    ind_gold = tup
                if current[tup][0] == w_date_silver:
                    ind_silver = (tup - 1)
                elif current[tup][0].timetuple().tm_mon == 1 and len(current) == 1:
                    ind_silver = tup
        return ind_gold, ind_silver
            
    def add_row(goldorsilver,add_row,w_date,w_value):
        '''Adds row with provided data to indicated xlsx file'''
        if goldorsilver == 'gold':
            add_ws = gold_ws
        elif goldorsilver == 'silver':
            add_ws = silver_ws
        add_ws.cell(row=add_row, column=1).value = w_date
        add_ws.cell(row=add_row, column=1).number_format = 'yy/mm/dd'
        add_ws.cell(row=add_row, column=2).value = float(w_value)
        add_ws.cell(row=add_row, column=2).number_format = '#,##0.00'
        add_ws.cell(row=add_row, column=3).value = '=(B{}-B{})'.format(add_row,add_row-1)
        add_ws.cell(row=add_row, column=3).number_format = '#,##0.00'
        add_ws.cell(row=add_row, column=4).value = '=AVERAGE(B{}:B{})'.format(add_row,add_row-199)
        add_ws.cell(row=add_row, column=4).number_format = '#,##0.00'
        add_ws.cell(row=add_row, column=5).value = '=(B{}/D{})'.format(add_row, add_row)
        add_ws.cell(row=add_row, column=5).number_format = '#,##0.00'

    gold_silver_index = (find_index())
    
    # Gold
    g_add_row = gold_mr + 1        
    for i in range(gold_silver_index[0],-1,-1):
        if not current[i][2] == '-':
            add_row('gold',g_add_row,current[i][0],current[i][2])
            g_add_row += 1
        elif current[i][2] == '-' and not current[i][1] == '-':
            add_row('gold',g_add_row,current[i][0],current[i][1])
            g_add_row += 1
        
    # Silver
    s_add_row = silver_mr + 1
    for i in range(gold_silver_index[1],-1,-1):
        if not current[i][3] == '-':
            add_row('silver',s_add_row,current[i][0],current[i][3])
            s_add_row += 1
            
    gold_wb.save(prices_gold)
    silver_wb.save(prices_silver)


def calc_dma():
    '''
    Retrieves and calculates the data for 200dma and relative gold and relative 
    silver values. This is also calculated in the xlsx file if loaded in a client,
    but openpyxl will not evaluate formulas so must be separately calculated in
    the script for alerts, etc.
    Output: tuple of (gold_200dma, gold_ratio, silver_200dma, silver_ratio)
    '''
    
    # Gold
    gold_200d = []
    for i in range(200):
        row = gold_mr - i
        gold_200d.append(gold_ws.cell(row=row, column=2).value)
    gold_200dma = (sum(gold_200d) / len(gold_200d))
    gold_ratio = (gold_ws.cell(row=gold_mr, column=2).value)/gold_200dma
    
    # Silver
    silver_200d = []
    for i in range(200):
        row = silver_mr - i
        silver_200d.append(silver_ws.cell(row=row, column=2).value)
    silver_200dma = (sum(silver_200d) / len(silver_200d))
    silver_ratio = (silver_ws.cell(row=silver_mr, column=2).value)/silver_200dma
    
    return gold_200dma, gold_ratio, silver_200dma, silver_ratio


def check_alerts(dmas):
    '''
    Checks for high and low R-Gold and R-Silver according to variables set globally.
    Input: tuple of (gold_200dma, gold_ratio, silver_200dma, silver_ratio)
    Output: tuple of (number of alerts, string of gold alert, string of silver alert)
    '''
    
    alerts = 0
    
    gold_last_date = gold_ws.cell(row=gold_mr, column=1).value.strftime('%Y-%m-%d')
    gold_last_price = gold_ws.cell(row=gold_mr, column=2).value
    silver_last_date = silver_ws.cell(row=silver_mr, column=1).value.strftime('%Y-%m-%d')
    silver_last_price = silver_ws.cell(row=silver_mr, column=2).value
    
    # Gold
    if  dmas[1] > rgold_low_alert and dmas[1] < rgold_high_alert:
        gold_alerts = 'Last Gold date: {}\nNo Gold alerts. Last Gold Price: ${:.2f} Gold 200dma: ${:.2f}  Relative Gold: {:.2f}'.format(gold_last_date,gold_last_price,dmas[0],dmas[1])
    elif dmas[1] <= rgold_low_alert:
        gold_alerts = 'Last Gold date: {}\nLOW Relative Gold. Check for buying conditions! Last Gold Price: ${:.2f} Gold 200dma: ${:.2f}  Relative Gold: {:.2f}'.format(gold_last_date,gold_last_price,dmas[0],dmas[1])
        alerts += 1
    elif dmas[1] >= rgold_high_alert:
        gold_alerts = 'Last Gold date: {}\nHIGH Relative Gold. Last Gold Price: ${:.2f} Check for selling conditions! Gold 200dma: ${:.2f}  Relative Gold: {:.2f}'.format(gold_last_date,gold_last_price,dmas[0],dmas[1])
        alerts += 1
    # Silver
    if  dmas[3] > rsilver_low_alert and dmas[3] < rsilver_high_alert:
        silver_alerts = 'Last Silver date: {}\nNo Silver alerts. Last Silver Price: ${:.2f} Silver 200dma: ${:.2f}  Relative Silver: {:.2f}'.format(silver_last_date,silver_last_price,dmas[2],dmas[3])
    elif dmas[3] <= rsilver_low_alert:
        silver_alerts = 'Last Silver date: {}\nLOW Relative Silver. Last Silver Price: ${:.2f} Check for buying conditions! Silver 200dma: ${:.2f}  Relative Silver: {:.2f}'.format(silver_last_date,silver_last_price,dmas[2],dmas[3])
        alerts += 1
    elif dmas[3] >= rsilver_high_alert:
        silver_alerts = 'Last Silver date: {}\nHIGH Relative Silver. Last Silver Price: ${:.2f} Check for selling conditions! Silver 200dma: ${:.2f}  Relative Silver: {:.2f}'.format(silver_last_date,silver_last_price,dmas[2],dmas[3])
        alerts += 1
    
    return alerts, gold_alerts, silver_alerts


def log_alerts():
    '''
    Logs day's results to text file.
    '''
    
    with open('goldsilver_alerts_log.txt', 'a') as alerts_log:
        alerts_log.write(str(datetime.datetime.now()) + '\n\n')
        alerts_log.write('Alerts: {}'.format(alerts[0]) + '\n')
        alerts_log.write(alerts[1] + '\n')
        alerts_log.write(alerts[2] + '\n\n')


# Global Variables
        
prices_gold = 'goldprices.xlsx'
prices_silver = 'silverprices.xlsx'
rgold_low_alert = 0.975
rgold_high_alert = 1.225
rsilver_low_alert = 0.8
rsilver_high_alert = 1.4


# Verify that xlsx files exist. Separate check in load_xlsx() for sufficient data.

if not path.exists(prices_gold) or not path.exists(prices_silver):
    raise MissingFiles('Error, missing data. Please ensure goldprices.xlsx and silverprices.xlsx have been generated appropriately.')


if __name__ == '__main__':
    current = get_current()
    load_xlsx()
    gold_mr = verify_mr('gold')
    silver_mr = verify_mr('silver')
    update()
    dmas = calc_dma()
    alerts = check_alerts(dmas)
    log_alerts()
    send_email(alerts)
#    if alerts[0] > 0:
#        send_email(alerts)

    
    # Function testing, uncomment to run functions with test data.
#    current = [(datetime.date(2020, 4, 20), '1684.95', '1686.20', '15.15'), (datetime.date(2020, 4, 17), '1693.15', '1692.55', '15.16'), (datetime.date(2020, 4, 16), '1717.85', '1729.50', '15.50'), (datetime.date(2020, 4, 15), '1712.25', '1718.65', '15.57'), (datetime.date(2020, 4, 14), '1715.85', '1741.90', '15.51'), (datetime.date(2020, 4, 9), '1662.50', '1680.65', '15.18'), (datetime.date(2020, 4, 8), '1649.05', '1647.80', '15.06'), (datetime.date(2020, 4, 7), '1652.20', '1649.25', '15.08'), (datetime.date(2020, 4, 6), '1636.60', '-', '-'), (datetime.date(2020, 4, 3), '1609.75', '1613.10', '14.39'), (datetime.date(2020, 4, 2), '1588.05', '-', '-'), (datetime.date(2020, 4, 1), '1594.25', '1576.55', '14.02'), (datetime.date(2020, 3, 31), '1604.65', '1608.95', '13.93'), (datetime.date(2020, 3, 30), '1624.45', '1618.30', '14.06'), (datetime.date(2020, 3, 27), '1621.20', '1617.30', '14.32'), (datetime.date(2020, 3, 26), '1620.10', '1634.80', '14.42'), (datetime.date(2020, 3, 25), '1620.95', '1605.45', '13.96'), (datetime.date(2020, 3, 24), '1599.50', '1605.75', '13.62'), (datetime.date(2020, 3, 23), '1494.50', '1525.40', '12.51'), (datetime.date(2020, 3, 20), '1504.45', '1494.40', '12.63'), (datetime.date(2020, 3, 19), '1480.70', '1474.25', '12.00'), (datetime.date(2020, 3, 18), '1506.00', '1498.20', '12.42'), (datetime.date(2020, 3, 17), '1472.35', '1536.20', '12.44'), (datetime.date(2020, 3, 16), '1504.65', '1487.70', '12.96'), (datetime.date(2020, 3, 13), '1588.15', '1562.80', '15.77'), (datetime.date(2020, 3, 12), '1636.65', '1570.70', '16.52'), (datetime.date(2020, 3, 11), '1662.50', '1653.75', '17.02'), (datetime.date(2020, 3, 10), '1657.40', '1655.70', '17.07'), (datetime.date(2020, 3, 9), '1676.60', '1672.50', '16.88'), (datetime.date(2020, 3, 6), '1687.00', '1683.65', '17.48'), (datetime.date(2020, 3, 5), '1647.45', '1659.60', '17.20'), (datetime.date(2020, 3, 4), '1644.80', '1641.85', '17.25'), (datetime.date(2020, 3, 3), '1599.05', '1615.50', '16.81'), (datetime.date(2020, 3, 2), '1609.70', '1599.65', '16.92'), (datetime.date(2020, 2, 28), '1626.35', '1609.85', '17.18'), (datetime.date(2020, 2, 27), '1646.60', '1652.00', '18.05'), (datetime.date(2020, 2, 26), '1647.95', '1634.90', '18.08'), (datetime.date(2020, 2, 25), '1655.90', '1650.30', '18.33'), (datetime.date(2020, 2, 24), '1682.35', '1671.65', '18.78'), (datetime.date(2020, 2, 21), '1633.70', '1643.30', '18.56'), (datetime.date(2020, 2, 20), '1610.35', '1619.00', '18.38'), (datetime.date(2020, 2, 19), '1609.50', '1604.20', '18.34'), (datetime.date(2020, 2, 18), '1588.20', '1589.85', '17.88'), (datetime.date(2020, 2, 17), '1580.30', '1580.80', '17.80'), (datetime.date(2020, 2, 14), '1576.35', '1581.40', '17.70'), (datetime.date(2020, 2, 13), '1575.00', '1575.05', '17.64'), (datetime.date(2020, 2, 12), '1566.75', '1563.70', '17.56'), (datetime.date(2020, 2, 11), '-', '1570.50', '17.70'), (datetime.date(2020, 2, 7), '1568.30', '1572.65', '17.77'), (datetime.date(2020, 2, 6), '1564.75', '1563.30', '17.76'), (datetime.date(2020, 2, 5), '1552.20', '1553.30', '17.62'), (datetime.date(2020, 2, 4), '1571.20', '1558.35', '17.73'), (datetime.date(2020, 2, 3), '1578.85', '1574.75', '17.77'), (datetime.date(2020, 1, 31), '1580.85', '1584.20', '17.88'), (datetime.date(2020, 1, 30), '1580.40', '1578.25', '17.72'), (datetime.date(2020, 1, 29), '1571.20', '-', '-'), (datetime.date(2020, 1, 28), '1579.60', '1574.00', '17.98'), (datetime.date(2020, 1, 27), '1583.45', '1580.10', '18.30'), (datetime.date(2020, 1, 24), '1561.85', '1564.30', '17.83'), (datetime.date(2020, 1, 23), '1554.05', '-', '-'), (datetime.date(2020, 1, 22), '1558.10', '1556.90', '17.77'), (datetime.date(2020, 1, 21), '1556.25', '1551.30', '17.98'), (datetime.date(2020, 1, 20), '1559.25', '1560.15', '18.06'), (datetime.date(2020, 1, 17), '1556.50', '1557.60', '18.06'), (datetime.date(2020, 1, 16), '1555.20', '1554.55', '18.01'), (datetime.date(2020, 1, 15), '1551.90', '1549.00', '17.85'), (datetime.date(2020, 1, 14), '1544.95', '1545.10', '17.77'), (datetime.date(2020, 1, 13), '1550.35', '1549.90', '17.98'), (datetime.date(2020, 1, 10), '1548.80', '1553.60', '17.92'), (datetime.date(2020, 1, 9), '1547.85', '1550.75', '17.91'), (datetime.date(2020, 1, 8), '1582.85', '1571.95', '18.42'), (datetime.date(2020, 1, 7), '1566.50', '1567.85', '18.14'), (datetime.date(2020, 1, 6), '1576.85', '1573.10', '18.44'), (datetime.date(2020, 1, 3), '1547.40', '1548.75', '18.21'), (datetime.date(2020, 1, 2), '1520.55', '1527.10', '17.92')]    
#    dmas = (1524.63, 1.11, 16.93, 0.89)
#    test_alerts = (0, 'No Gold alerts. Gold 200dma: $1524.63  Relative Gold: 1.11', 'No Silver alerts. Silver 200dma: $16.93  Relative Silver: 0.89')
#    send_email.send_alert(test_alerts)
